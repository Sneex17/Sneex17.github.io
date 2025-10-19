import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
import urllib.parse

# Archivos de Excel
archivos = {
    "Familias": "Familias.xlsx",
    "Personas1": "Personales.xlsx",
    "Personas2": "Persona1 y Persona2.xlsx"  # Cambia el nombre según tu tercer Excel
}


# ------------------- Función para actualizar enlaces -------------------
def actualizar_excel(tipo, archivo):
    df = pd.read_excel(archivo)

    for index, row in df.iterrows():
        if pd.isna(row['Enlaces']) or row['Enlaces'] == "":
            if tipo == "Familias":
                fam1 = str(row['Familia1']).strip() if not pd.isna(row['Familia1']) else ""
                fam2 = str(row['Familia2']).strip() if not pd.isna(row['Familia2']) else ""
                cantidad = row['Cantidad']

                enlace = f"https://sneex17.github.io/#Familia1={urllib.parse.quote(fam1)}&Familia2={urllib.parse.quote(fam2)}&Cantidad={cantidad}"
                print(f"Familia: {fam1} {fam2} | enlace: {enlace}")
                df.at[index, 'Enlaces'] = enlace

            else:  # Personas
                persona1 = str(row['Nombre']).strip() if not pd.isna(row['Nombre']) else ""
                persona2 = str(row['Apellido']).strip() if not pd.isna(row['Apellido']) else ""
                cantidad = row['Cantidad']

                if cantidad > 1:
                    enlace = f"https://sneex17.github.io/#Persona1={urllib.parse.quote(persona1)}&Persona2={urllib.parse.quote(persona2)}&Cantidad={cantidad}"
                    print(f"Personas: {persona1} {persona2} | enlace: {enlace}")
                else:
                    enlace = f"https://sneex17.github.io/#Nombre={urllib.parse.quote(persona1)}&Apellido={urllib.parse.quote(persona2)}&Cantidad={cantidad}"
                    print(f"Persona: {persona1} {persona2} | enlace: {enlace}")
                df.at[index, 'Enlaces'] = enlace

    # Guardar y aplicar fuente Times New Roman 11
    df.to_excel(archivo, index=False)
    wb = load_workbook(archivo)
    ws = wb.active
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Times New Roman', size=11)
    wb.save(archivo)


# ------------------- Procesar todos los archivos -------------------
actualizar_excel("Familias", archivos["Familias"])
actualizar_excel("Personas", archivos["Personas1"])
actualizar_excel("Personas", archivos["Personas2"])

print("\n✅ Todos los archivos actualizados con enlaces y fuente Times New Roman 11.")
